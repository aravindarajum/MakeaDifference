import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By

workbook_name = 'C://Users//Lenovo//PycharmProjects//MakeADifference//testdata//UrbanSustainability_testdata.xlsx'
sheet_name = 'Sheet1'
driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://aravindaec.wixsite.com/makeadifference/events/panel-paving-the-way-urban-sustainability/form")


class readinglistfromexcel:
    def selectcheckboxes(self):
        hobbies = self.read_data_from_excel(self, workbook_name)
        for hobby in hobbies:
            print(hobby)
            if hobby == 'music':
                driver.find_element(By.XPATH, "//div[text()='Music']").click()
            if hobby == 'chess':
                driver.find_element(By.XPATH, "//div[text()='Chess']").click()
            if hobby == 'crafts':
                driver.find_element(By.XPATH, "//div[text()='Crafts']").click()

    def read_data_from_excel(self, workbook_n, sheet_n):
        workbook = openpyxl.load_workbook(workbook_n)
        sheet = workbook[sheet_n]
        data = []
        data = [sheet.cell(row=i, column=7).value for i in range(2, 5)]
        print(data)
        # for row in sheet.iter_rows(min_row=2, col="G"):
        #     data.append(row)
        return data


classobject = readinglistfromexcel()
classobject.read_data_from_excel(workbook_name,sheet_name)
classobject.selectcheckboxes()